"""
Builds a single combined CSV/Excel report from captioning results and
(optionally) validation results, including a per-label True/False
Positive/Negative verdict and overall Precision/Recall/F1/Accuracy.

When include_validation is False, the validation columns, verdicts, and
metrics summary are omitted entirely - since validation never ran and
there's nothing to show.

--- Why these particular metrics ---
GROUND_TRUTH only lists labels that ARE present in each video - it never
states a label is definitely absent. That means there is no true "negative"
class to draw True Negatives from, so classic Accuracy (which needs TNs)
isn't strictly meaningful here. Instead:

  - Object detection (a located bounding box) is treated as the ground-truth
    signal for "is this label really there".
  - The caption TEXT is the prediction being evaluated against it.

Per (video, label) instance:
  True Positive   - detected AND mentioned in caption      -> correct
  False Negative  - detected but NOT mentioned in caption  -> caption missed it
  False Positive  - NOT detected but caption claims it     -> caption hallucinated

Precision and Recall follow the standard formulas from these counts.
"Accuracy" is reported as TP / (TP + FP + FN) (no TN term exists) - this is
the IoU-style accuracy used in detection tasks that lack true negatives,
and is labeled as such rather than implying textbook accuracy.
"""

import csv
import os
from typing import Dict, List, Tuple

from models import CaptionRecord, ValidationResult


def _format_labels(labels: List[str], flags: dict) -> str:
    return ", ".join(f"{label} ({'Yes' if flags.get(label) else 'No'})" for label in labels)


def _classify_label(caption_match: bool, detection_match: bool) -> str:
    """Per-label verdict. Detection is treated as ground truth; the caption
    text is the prediction being evaluated against it."""
    if detection_match and caption_match:
        return "True Positive"
    if detection_match and not caption_match:
        return "False Negative"
    if not detection_match and caption_match:
        return "False Positive"
    return "False Negative"  # neither signal found a label that should be present


def _format_verdicts(labels: List[str], caption_matches: dict, detection_matches: dict) -> str:
    parts = []
    for label in labels:
        verdict = _classify_label(caption_matches.get(label, False), detection_matches.get(label, False))
        parts.append(f"{label}: {verdict}")
    return "; ".join(parts)


class MetricsCalculator:
    """Tallies per-label verdicts across all videos and computes
    Precision, Recall, F1, and IoU-style Accuracy."""

    def __init__(self, records: List[CaptionRecord], validation_results: Dict[str, ValidationResult]):
        self.tp = 0
        self.fp = 0
        self.fn = 0

        for r in records:
            v = validation_results.get(r.video)
            if not v:
                continue
            for label in v.expected_labels:
                verdict = _classify_label(
                    v.caption_matches.get(label, False),
                    v.detection_matches.get(label, False),
                )
                if verdict == "True Positive":
                    self.tp += 1
                elif verdict == "False Positive":
                    self.fp += 1
                elif verdict == "False Negative":
                    self.fn += 1

    def precision(self) -> float:
        denom = self.tp + self.fp
        return self.tp / denom if denom else 0.0

    def recall(self) -> float:
        denom = self.tp + self.fn
        return self.tp / denom if denom else 0.0

    def f1(self) -> float:
        p, r = self.precision(), self.recall()
        return (2 * p * r / (p + r)) if (p + r) else 0.0

    def accuracy(self) -> float:
        """IoU-style accuracy: TP / (TP + FP + FN). No True Negative term
        exists in this setup (see module docstring)."""
        denom = self.tp + self.fp + self.fn
        return self.tp / denom if denom else 0.0

    def as_rows(self) -> List[Tuple[str, str]]:
        return [
            ("True Positives", str(self.tp)),
            ("False Positives", str(self.fp)),
            ("False Negatives", str(self.fn)),
            ("Precision", f"{self.precision():.3f}"),
            ("Recall", f"{self.recall():.3f}"),
            ("F1 Score", f"{self.f1():.3f}"),
            ("Accuracy (TP / (TP+FP+FN), no true negatives in this setup)", f"{self.accuracy():.3f}"),
        ]


class CombinedReportBuilder:
    def __init__(
        self,
        records: List[CaptionRecord],
        validation_results: Dict[str, ValidationResult],
        include_validation: bool,
    ):
        self.records = records
        self.validation_results = validation_results
        self.include_validation = include_validation
        self.max_frames = max((len(r.frame_captions) for r in records), default=0)
        self.metrics = MetricsCalculator(records, validation_results) if include_validation else None

    def build_csv(self, path: str) -> None:
        frame_headers = [f"Frame {i + 1} Caption" for i in range(self.max_frames)]
        headers = ["Video", "Detailed Analysis", "Summary Caption"] + frame_headers

        if self.include_validation:
            headers += ["Expected Labels", "Classification",
                        "Caption Matches", "Detection Matches", "Verdict", "Notes"]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for r in self.records:
                row = [r.video, r.detailed_caption, r.summary_caption]
                row += r.frame_captions + [""] * (self.max_frames - len(r.frame_captions))

                if self.include_validation:
                    v = self.validation_results.get(r.video)
                    if v:
                        row += [
                            ", ".join(v.expected_labels),
                            v.classification,
                            _format_labels(v.expected_labels, v.caption_matches),
                            _format_labels(v.expected_labels, v.detection_matches),
                            _format_verdicts(v.expected_labels, v.caption_matches, v.detection_matches),
                            v.notes,
                        ]
                    else:
                        row += ["", "", "", "", "", ""]

                writer.writerow(row)

            if self.include_validation and self.metrics:
                writer.writerow([])
                writer.writerow(["--- Summary Metrics ---"])
                for label, value in self.metrics.as_rows():
                    writer.writerow([label, value])

    def build_excel(self, path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Video Caption & Validation Report"

        frame_headers = [f"Frame {i + 1} Caption" for i in range(self.max_frames)]
        headers = ["#", "Video", "Detailed Analysis", "Summary Caption"] + frame_headers

        if self.include_validation:
            headers += ["Annotated Frame", "Expected Labels", "Classification",
                        "Caption Matches", "Detection Matches", "Verdict", "Notes"]

        ws.append(headers)

        header_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        all_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        partial_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
        none_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

        widths = [4, 30, 45, 40] + [30] * self.max_frames
        if self.include_validation:
            widths += [26, 22, 18, 28, 28, 34, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for idx, r in enumerate(self.records, start=2):
            ws.row_dimensions[idx].height = 110
            col = 1

            ws.cell(row=idx, column=col, value=idx - 1)
            col += 1
            ws.cell(row=idx, column=col, value=r.video).alignment = Alignment(wrap_text=True, vertical="center")
            col += 1
            ws.cell(row=idx, column=col, value=r.detailed_caption).alignment = Alignment(wrap_text=True, vertical="center")
            col += 1
            ws.cell(row=idx, column=col, value=r.summary_caption).alignment = Alignment(wrap_text=True, vertical="center")
            col += 1

            for i in range(self.max_frames):
                val = r.frame_captions[i] if i < len(r.frame_captions) else ""
                ws.cell(row=idx, column=col, value=val).alignment = Alignment(wrap_text=True, vertical="center")
                col += 1

            if self.include_validation:
                v = self.validation_results.get(r.video)
                img_col_letter = get_column_letter(col)
                col += 1

                if v:
                    ws.cell(row=idx, column=col, value=", ".join(v.expected_labels)).alignment = Alignment(
                        wrap_text=True, vertical="center")
                    col += 1

                    cls_cell = ws.cell(row=idx, column=col, value=v.classification)
                    cls_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if v.classification.startswith("All Matched"):
                        cls_cell.fill = all_fill
                    elif v.classification.startswith("Partial"):
                        cls_cell.fill = partial_fill
                    elif v.classification.startswith("None"):
                        cls_cell.fill = none_fill
                    col += 1

                    ws.cell(row=idx, column=col, value=_format_labels(v.expected_labels, v.caption_matches)).alignment = \
                        Alignment(wrap_text=True, vertical="center")
                    col += 1
                    ws.cell(row=idx, column=col, value=_format_labels(v.expected_labels, v.detection_matches)).alignment = \
                        Alignment(wrap_text=True, vertical="center")
                    col += 1
                    ws.cell(row=idx, column=col,
                            value=_format_verdicts(v.expected_labels, v.caption_matches, v.detection_matches)
                            ).alignment = Alignment(wrap_text=True, vertical="center")
                    col += 1
                    ws.cell(row=idx, column=col, value=v.notes).alignment = Alignment(wrap_text=True, vertical="center")
                    col += 1

                    if v.annotated_image_path and os.path.exists(v.annotated_image_path):
                        img = XLImage(v.annotated_image_path)
                        img.width = 170
                        img.height = 120
                        ws.add_image(img, f"{img_col_letter}{idx}")

        if self.include_validation and self.metrics:
            self._write_metrics_sheet(wb)

        wb.save(path)

    def _write_metrics_sheet(self, wb) -> None:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Summary Metrics")
        ws.column_dimensions["A"].width = 55
        ws.column_dimensions["B"].width = 15

        title_font = Font(bold=True, size=13, color="2E4053")
        ws.cell(row=1, column=1, value="Overall Model Accuracy Metrics").font = title_font

        header_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.cell(row=3, column=1, value="Metric").fill = header_fill
        ws.cell(row=3, column=1).font = header_font
        ws.cell(row=3, column=2, value="Value").fill = header_fill
        ws.cell(row=3, column=2).font = header_font

        for i, (label, value) in enumerate(self.metrics.as_rows(), start=4):
            ws.cell(row=i, column=1, value=label).alignment = Alignment(wrap_text=True)
            ws.cell(row=i, column=2, value=value).alignment = Alignment(horizontal="center")

        note_row = 4 + len(self.metrics.as_rows()) + 2
        note = (
            "Note: True Negatives do not exist in this setup, since GROUND_TRUTH only "
            "lists labels expected to be PRESENT (never explicitly absent). Object "
            "detection is treated as the ground-truth signal; caption text is the "
            "prediction being evaluated against it. Accuracy above is TP/(TP+FP+FN), "
            "the standard substitute used in detection tasks lacking a true negative class."
        )
        ws.cell(row=note_row, column=1, value=note).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 3, end_column=4)