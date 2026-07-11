"""Builds the final CSV and Excel deliverables from a list of ValidationResult."""

import csv
import os
from typing import List

from models import ValidationResult


class ReportBuilder:
    def __init__(self, results: List[ValidationResult]):
        self.results = results

    def build_csv(self, path: str) -> None:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "Video", "Expected Label", "Classification",
                "Caption Match", "Matched Term", "Detection Match", "Notes",
            ])
            for r in self.results:
                writer.writerow([
                    r.video, r.expected_label, r.classification,
                    r.caption_match, r.caption_matched_term, r.detection_match, r.notes,
                ])

    def build_excel(self, path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Report"

        headers = ["#", "Video", "Annotated Frame", "Expected Label",
                   "Classification", "Caption Match", "Detection Match", "Notes"]
        ws.append(headers)

        header_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        tp_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        fp_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        fn_fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

        col_widths = [4, 32, 26, 16, 30, 14, 14, 35]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for idx, r in enumerate(self.results, start=2):
            ws.row_dimensions[idx].height = 110
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=r.video).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row=idx, column=4, value=r.expected_label).alignment = Alignment(vertical="center")

            cls_cell = ws.cell(row=idx, column=5, value=r.classification)
            cls_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if r.classification == "True Positive":
                cls_cell.fill = tp_fill
            elif "False Positive" in r.classification:
                cls_cell.fill = fp_fill
            elif "False Negative" in r.classification:
                cls_cell.fill = fn_fill

            ws.cell(row=idx, column=6, value=str(r.caption_match)).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=idx, column=7, value=str(r.detection_match)).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=idx, column=8, value=r.notes).alignment = Alignment(wrap_text=True, vertical="center")

            if r.annotated_image_path and os.path.exists(r.annotated_image_path):
                img = XLImage(r.annotated_image_path)
                img.width = 170
                img.height = 120
                ws.add_image(img, f"C{idx}")

        wb.save(path)
