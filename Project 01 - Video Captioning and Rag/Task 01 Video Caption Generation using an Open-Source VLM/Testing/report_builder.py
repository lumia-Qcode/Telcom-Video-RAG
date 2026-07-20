"""
Reads/writes the combined video report as plain dict rows - no classes.

The CSV is the source of truth for "what's already been processed", which
is what makes month-end runs incremental: run_pipeline.py reads it back in,
skips any video already listed, processes only the new ones, and rewrites
both the CSV and the Excel from the full (old + new) row list.
"""

import csv
import os

FIELDNAMES = [
    "video", "detailed_caption", "summary_caption", "frame_captions",
    "claimed_entities", "confirmed_entities", "classification",
    "hallucination_rate", "risk_band", "annotated_image_path",
]

FRAME_SEP = " | "
ENTITY_SEP = ", "


def row_to_csv_dict(row):
    """Flattens a process_video() dict (which has real lists) into strings
    for CSV storage."""
    return {
        "video": row["video"],
        "detailed_caption": row["detailed_caption"],
        "summary_caption": row["summary_caption"],
        "frame_captions": FRAME_SEP.join(row["frame_captions"]),
        "claimed_entities": ENTITY_SEP.join(row["claimed_entities"]),
        "confirmed_entities": row["confirmed_entities"],
        "classification": row["classification"],
        "hallucination_rate": f"{row['hallucination_rate']:.4f}",
        "risk_band": row["risk_band"],
        "annotated_image_path": row["annotated_image_path"],
    }


def csv_dict_to_row(d):
    """Reverses row_to_csv_dict() when reading an existing report back in."""
    return {
        "video": d["video"],
        "detailed_caption": d["detailed_caption"],
        "summary_caption": d["summary_caption"],
        "frame_captions": d["frame_captions"].split(FRAME_SEP) if d["frame_captions"] else [],
        "claimed_entities": d["claimed_entities"].split(ENTITY_SEP) if d["claimed_entities"] else [],
        "confirmed_entities": d["confirmed_entities"],
        "classification": d["classification"],
        "hallucination_rate": float(d["hallucination_rate"]) if d["hallucination_rate"] else 0.0,
        "risk_band": d["risk_band"],
        "annotated_image_path": d["annotated_image_path"],
    }


def read_existing_rows(path):
    """Returns (rows, processed_video_names). Empty if the report doesn't
    exist yet (first run)."""
    if not os.path.exists(path):
        return [], set()

    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for d in reader:
            # Stop at the blank line that starts the summary block, if present.
            if not d.get("video"):
                break
            rows.append(csv_dict_to_row(d))

    processed = {r["video"] for r in rows}
    return rows, processed


def dataset_summary(rows):
    classified = [r for r in rows if r["claimed_entities"]]
    total_videos = len(rows)
    classified_videos = len(classified)
    avg_rate = (sum(r["hallucination_rate"] for r in classified) / len(classified)) if classified else 0.0

    band_counts = {}
    for r in classified:
        band_counts[r["risk_band"]] = band_counts.get(r["risk_band"], 0) + 1

    summary_rows = [
        ("Total videos", str(total_videos)),
        ("Videos with extracted entities", str(classified_videos)),
        ("Average hallucination rate", f"{avg_rate:.1%}"),
    ]
    for band in ["Low (0%)", "Moderate (1-25%)", "Elevated (26-50%)", "High (51-75%)", "Severe (76-100%)"]:
        summary_rows.append((f"Videos in risk band: {band}", str(band_counts.get(band, 0))))
    return summary_rows


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for r in rows:
            writer.writerow(row_to_csv_dict(r))

        writer.writerow({})
        f.write("--- Dataset Summary ---\n")
        for label, value in dataset_summary(rows):
            f.write(f"{label},{value}\n")


def write_excel(rows, path):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    max_frames = max((len(r["frame_captions"]) for r in rows), default=0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Caption Hallucination Report"

    frame_headers = [f"Frame {i + 1} Caption" for i in range(max_frames)]
    headers = (["#", "Video", "Detailed Analysis", "Summary Caption"] + frame_headers +
               ["Annotated Frame", "Claimed Entities", "Confirmed Entities",
                "Classification", "Hallucination Rate", "Risk Band"])
    ws.append(headers)

    header_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    band_fills = {
        "Low (0%)": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        "Moderate (1-25%)": PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
        "Elevated (26-50%)": PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid"),
        "High (51-75%)": PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid"),
        "Severe (76-100%)": PatternFill(start_color="F1948A", end_color="F1948A", fill_type="solid"),
    }

    widths = [4, 30, 45, 40] + [30] * max_frames + [26, 30, 30, 20, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for idx, r in enumerate(rows, start=2):
        ws.row_dimensions[idx].height = 110
        col = 1
        ws.cell(row=idx, column=col, value=idx - 1); col += 1
        ws.cell(row=idx, column=col, value=r["video"]).alignment = Alignment(wrap_text=True, vertical="center"); col += 1
        ws.cell(row=idx, column=col, value=r["detailed_caption"]).alignment = Alignment(wrap_text=True, vertical="center"); col += 1
        ws.cell(row=idx, column=col, value=r["summary_caption"]).alignment = Alignment(wrap_text=True, vertical="center"); col += 1

        for i in range(max_frames):
            val = r["frame_captions"][i] if i < len(r["frame_captions"]) else ""
            ws.cell(row=idx, column=col, value=val).alignment = Alignment(wrap_text=True, vertical="center")
            col += 1

        img_col_letter = get_column_letter(col)
        col += 1

        ws.cell(row=idx, column=col, value=", ".join(r["claimed_entities"])).alignment = \
            Alignment(wrap_text=True, vertical="center"); col += 1
        ws.cell(row=idx, column=col, value=r["confirmed_entities"]).alignment = \
            Alignment(wrap_text=True, vertical="center"); col += 1
        ws.cell(row=idx, column=col, value=r["classification"]).alignment = \
            Alignment(wrap_text=True, vertical="center"); col += 1
        ws.cell(row=idx, column=col, value=f"{r['hallucination_rate']:.1%}" if r["claimed_entities"] else "").alignment = \
            Alignment(horizontal="center", vertical="center"); col += 1
        band = r["risk_band"]
        band_cell = ws.cell(row=idx, column=col, value=band)
        band_cell.alignment = Alignment(horizontal="center", vertical="center")
        if band in band_fills:
            band_cell.fill = band_fills[band]
        col += 1

        if r["annotated_image_path"] and os.path.exists(r["annotated_image_path"]):
            img = XLImage(r["annotated_image_path"])
            img.width, img.height = 170, 120
            ws.add_image(img, f"{img_col_letter}{idx}")

    _write_summary_sheet(wb, rows)
    wb.save(path)


def _write_summary_sheet(wb, rows):
    from openpyxl.styles import Font, Alignment, PatternFill

    ws = wb.create_sheet("Dataset Summary")
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 15

    ws.cell(row=1, column=1, value="Caption Hallucination Summary (no manual ground truth used)").font = \
        Font(bold=True, size=13, color="2E4053")

    header_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
    ws.cell(row=3, column=1, value="Metric").fill = header_fill
    ws.cell(row=3, column=1).font = Font(color="FFFFFF", bold=True)
    ws.cell(row=3, column=2, value="Value").fill = header_fill
    ws.cell(row=3, column=2).font = Font(color="FFFFFF", bold=True)

    summary_rows = dataset_summary(rows)
    for i, (label, value) in enumerate(summary_rows, start=4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value).alignment = Alignment(horizontal="center")

    note_row = 4 + len(summary_rows) + 2
    note = (
        "Hallucination rate = share of the caption's OWN claimed entities that the "
        "object detector could not locate in any frame. Entities are extracted "
        "automatically from each caption - there is no manually maintained ground "
        "truth list, so new videos need no extra setup. Run build_review_sample.py "
        "to pull a prioritized, human-reviewable sample weighted toward the "
        "highest-risk videos."
    )
    ws.cell(row=note_row, column=1, value=note).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 4, end_column=4)